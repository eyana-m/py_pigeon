
import numpy as np
import pandas as pd
from pandas import ExcelWriter
import os
import pathlib
from pathlib import Path
import io
import glob
import itertools
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import time
from time import sleep
import datetime
import csv

FILE_MASTERLIST = '/Users/eyanamallari/Projects/py_distribute_data/Input/CONTACTS_ALL_STATIC.xlsx'
FILE_TEMPLATE ='/Users/eyanamallari/Projects/py_distribute_data/Input/Contacts_Template.xlsx'

OUTPUT_DIRECTORY = '/Users/eyanamallari/Projects/py_distribute_data/Output/'
EXCLUDE_LIST = []
INCLUDE_LIST = ['Calista Rosales', 'Bianca Cardenas', 'Colette Black']

now = datetime.datetime.now()
QUARTER = "2018Q3"
MONTH = "Sept 2018"
VERSION = str(now.month).zfill(2) + str(now.day).zfill(2)

df = pd.read_excel(FILE_MASTERLIST,"CONTACTS_FINAL")


# ---------------------------------------
# Returns a list of Sales Representatives included in the split
# ---------------------------------------
def getSalesRep():
    df['Sales Representative'].fillna('Unknown', inplace = True)
    print('Getting all Sales Representatives')

    # --- EXCLUDE REPS ----
    #df_filtered = df[~df['Sales Representative'].isin(EXCLUDE_LIST)]
    #df_roster = df_filtered['Sales Representative'].unique()


    # --- INCLUDE REPS ----
    df_filtered = df[df['Sales Representative'].isin(INCLUDE_LIST)]
    df_roster = df_filtered['Sales Representative'].unique()

    #df_roster = df['Sales Representative'].unique()

    print(df_roster)
    d = df_roster.tolist()
    d_final = d

    #d_final = {k: d[k]for k in list(d)}

    print(d)
    print(len(d))

    print(d_final)
    print(len(d_final))
    return d_final



# ---------------------------------------
# Splits: Write Splits to Excel
# ---------------------------------------
def writeExcelFileByRep(owner_value, output_folder):


    owner = str(owner_value)

    # Filter by owner
    df_abridged = df[df['Sales Representative']==owner]
    rows_target = dataframe_to_rows(df_abridged)

    # ------------
    # Write to Excel
    # ------------

    FILE_PATH = output_folder
    print(FILE_PATH)


    book = load_workbook(FILE_PATH)
    writer = ExcelWriter(FILE_PATH, engine='openpyxl')

    writer.book = book

    for sheet in book.worksheets:
        if sheet.title == 'Contacts':

            for row in sheet['A1:H4']:
               for cell in row:
                   cell.value = None

            # Replenish
            for r_idx, row in enumerate(rows_target, 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)


    constant_tries = 2000
    tries = 2000

    assert tries > 0
    error = None
    result = None

    while tries:
        try:
            writer.save()
            writer.close()
        except IOError as e:
            error = e
            tries -= 1
            print('Attempt #', (constant_tries-tries)+1)
        except ValueError as e:
            error = e
            tries -= 1
            print('Attempt #', (constant_tries-tries)+1)
        else:
            break
    if not tries:
        print('Attempt #', (constant_tries-tries)+1)
        raise error

    print('Attempt #', (constant_tries-tries)+1)
    #print(df_abridged.loc[:,'Company':'Industry'].head(5))
    print("Done writing Excel file!")




# ---------------------------------------
# Creates files per included Sales Representative
# ---------------------------------------
def loopRosterCreateFiles(reps):

    print("Creating files!")
    print(reps)

    count = 1

    for rep in reps:
        print('------')

        print(count, rep)

        count = count + 1

        output_rep = "Contacts List - " + str(rep)
        output_folder = OUTPUT_DIRECTORY+ output_rep

        # Creates Folder for each rep
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        print("Created folder for", rep)

        # Set File Names
        rep_excel_file = QUARTER+ " Contact List - "+str(rep)+ " " + VERSION +".xlsx".strip()
        rep_excel_file_no_ext = QUARTER+ " Contact List - "+str(rep)+ " " + VERSION.strip()
        rep_excel_path = output_folder+"/"+rep_excel_file

        # Copies template
        shutil.copy(FILE_TEMPLATE,rep_excel_path)

        # Writes Filtered Data to Excel
        writeExcelFileByRep(rep,rep_excel_path)

        # Uploads to Google Drive
        #loopGSpreadsheet(writeToGDrive(rep_excel_file_no_ext,rep_excel_path,getFolder(rep)))


def main():

    loopRosterCreateFiles(getSalesRep())

if __name__ == '__main__':
    main()
