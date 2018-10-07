from __future__ import print_function
#googleapiclient.discovery
from apiclient.discovery  import build
from httplib2 import Http
from oauth2client import file, client, tools
from oauth2client.contrib import gce
from apiclient.http import MediaFileUpload
import numpy as np
import pandas as pd
from pandas import ExcelWriter
import os
import pathlib
from pathlib import Path
import io
import glob
import itertools
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)

import random
import time
from time import sleep
import datetime
import csv


CLIENT_SECRET = "/Users/eyanamallari/Projects/py_distribute_data/client_secret.json"



FILE_MASTERLIST = '/Users/eyanamallari/Projects/py_distribute_data/Input/CONTACTS_ALL_STATIC.xlsx'
FILE_TEMPLATE ='/Users/eyanamallari/Projects/py_distribute_data/Input/Contacts_Template.xlsx'
FILE_FOLDERS = '/Users/eyanamallari/Projects/py_distribute_data/Input/Pigeon Masterlist - Folders Lookup.csv'
OUTPUT_DIRECTORY = '/Users/eyanamallari/Projects/py_distribute_data/Output/'
EXCLUDE_LIST = []
INCLUDE_LIST = ['Calista Rosales', 'Bianca Cardenas', 'Colette Black']

now = datetime.datetime.now()
QUARTER = "2018Q3"
MONTH = "Sept 2018"
VERSION = str(now.month).zfill(2) + str(now.day).zfill(2)

df = pd.read_excel(FILE_MASTERLIST,"CONTACTS_FINAL")

# --------------------------------
# GDrive API: GDrive Authorization
# --------------------------------

SCOPES='https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets'
store = file.Storage('token.json')
creds = store.get()
if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets(CLIENT_SECRET, SCOPES)
    creds = tools.run_flow(flow, store)
SERVICE = build('drive', 'v3', http=creds.authorize(Http()))
SS_SERVICE = build('sheets', 'v4', http=creds.authorize(Http()))


PARENT_FOLDER = '19FBo4iSjyCS3NcqX6zaNgxtXWMqW7AyM'
# ------------------------------------
# GDrive API: Check if Filename exists
# ------------------------------------
def fileInGDrive(filename):
    results = SERVICE.files().list(q="mimeType='application/vnd.google-apps.spreadsheet' and name='"+filename+"' and trashed = false and parents in '"+PARENT_FOLDER+"'",fields="nextPageToken, files(id, name)").execute()
    items = results.get('files', [])
    if items:
        return True
    else:
        return False

# ------------------------------------
# GDrive API: Check if Folder exists
# ------------------------------------
def folderInGDrive(filename):
    results = SERVICE.files().list(q="mimeType='application/vnd.google-apps.folder' and name='"+filename+"' and trashed = false and parents in '"+PARENT_FOLDER+"'",fields="nextPageToken, files(id, name)").execute()
    items = results.get('files', [])
    if items:
        return True
    else:
        return False

# ---------------------------------------
# GDrive API: Create New Folder
# ---------------------------------------
def createGDriveFolder(filename,parent):
    file_metadata = {'name': filename,'parents': [parent],
    'mimeType': "application/vnd.google-apps.folder"}

    folder = SERVICE.files().create(body=file_metadata,
                                        fields='id').execute()
    print('Upload Success!')
    print('FolderID:', folder.get('id'))
    return folder.get('id')


# ---------------------------------------
# GDrive API: Upload files to Google Drive
# ---------------------------------------
def writeToGDrive(filename,source,folder_id):
    file_metadata = {'name': filename,'parents': [folder_id],
    'mimeType': 'application/vnd.google-apps.spreadsheet'}
    media = MediaFileUpload(source,
                            mimetype='application/vnd.ms-excel')

    if fileInGDrive(filename) is False:
        file = SERVICE.files().create(body=file_metadata,
                                            media_body=media,
                                            fields='id').execute()
        print('Upload Success!')
        print('File ID:', file.get('id'))
        return file.get('id')

    else:
        print('File already exists as', filename)


# ---------------------------------------
# GSheet API: Freeze Cells
# ---------------------------------------
def freezeCells(spreadsheet_id,sheet_id):

    my_range = {
        'sheetId': sheet_id
    }
    requests = [{
            "updateSheetProperties": {
                'properties': {
                    'sheetId': sheet_id,
                    'gridProperties': { 'frozenRowCount': 1,'frozenColumnCount': 2}
                },
                'fields': 'gridProperties(frozenRowCount,frozenColumnCount)'
            }
        }
    ]

    body = {
        'requests': requests
    }
    response = SS_SERVICE.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
    print('{0} update.'.format(len(response.get('replies'))));


# ---------------------------------------
# GSheet API: Protect Cells
# ---------------------------------------
def protectCells(spreadsheet_id,sheet_id):

    my_range = {
        'sheetId': sheet_id
    }
    requests = [{
      "addProtectedRange": {
        "protectedRange": {
          "range": {
                "sheetId": sheet_id,
                "startColumnIndex": 0,
                "endColumnIndex": 12,
            },
         "editors": {
            "users": [
              "eyanamallari@gmail.com"
            ]
          }

          }
     }
    }

    ]

    body = {
        'requests': requests
    }
    response = SS_SERVICE.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
    print('{0} update.'.format(len(response.get('replies'))));

# sheet_metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
# sheets = sheet_metadata.get('sheets', '')
# title = sheets[0].get("properties", {}).get("title", "Sheet1")
# sheet_id = sheets[0].get("properties", {}).get("sheetId", 0)


# ---------------------------------------
# GSheet API: Freeze Cells
# ---------------------------------------
def deleteCells(spreadsheet_id,sheet_id):

    my_range = {
        'sheetId': sheet_id
    }
    requests = [
    {
      "deleteDimension": {
        "range": {
          "sheetId": sheet_id,
          "dimension": "COLUMNS",
          "startIndex": 8,
          "endIndex": 26
        }
      }
    }
    ]

    body = {
        'requests': requests
    }
    response = SS_SERVICE.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
    print('{0} update.'.format(len(response.get('replies'))));

# ---------------------------------------
# GSheet API: Loop
# ---------------------------------------

def loopGSpreadsheet(spreadsheet_id):
    sheet_metadata = SS_SERVICE.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()

    sheets = sheet_metadata.get('sheets', '')

    for sheet in sheets:
        sheet_title = sheet.get("properties", {}).get("title")
        print("Processing ", sheet_title)
        #print(sheet)

        sheet_id = sheet.get("properties", {}).get("sheetId")
        freezeCells(spreadsheet_id,sheet_id)
        #protectCells(spreadsheet_id,sheet_id)

        column_count = sheet.get("properties", {}).get("gridProperties").get("columnCount")
        print(column_count, "columns")

        if column_count > 8:
            deleteCells(spreadsheet_id,sheet_id)




def getFolder(person):
    with open(FILE_FOLDERS, mode='r') as infile:
        reader = csv.reader(infile)
        folders_lookup = {rows[0]:rows[3]for rows in reader}

    person_string = str(person).strip()
    if person_string in folders_lookup:
        folder_id = folders_lookup.get(person_string)
        return folder_id
    else:
        return "No Folder"
        print('FOLDER NOT FOUND!!!!!', person)





# ---------------------------------------
# ---------------------------------------
# ---------------------------------------
# Part 1: Split files by User
# ---------------------------------------
# ---------------------------------------
# ---------------------------------------

# ---------------------------------------
# Returns a list of Sales Representatives included in the split
# ---------------------------------------
def getSalesRep():
    df['Sales Representative'].fillna('Unknown', inplace = True)
    print('Getting all Sales Representatives')
    # --- EXCLUDE REPS ----
    #df_filtered = df[~df['Sales Representative'].isin(EXCLUDE_LIST)]
    #df_roster = df_filtered['Sales Representative'].unique()


    # --- INCLUDE REPS ----
    df_filtered = df[df['Sales Representative'].isin(INCLUDE_LIST)]
    df_roster = df_filtered['Sales Representative'].unique()

    #df_roster = df['Sales Representative'].unique()

    print(df_roster)
    d = df_roster.tolist()
    d_final = d

    #d_final = {k: d[k]for k in list(d)}

    print(d)
    print(len(d))

    print(d_final)
    print(len(d_final))
    return d_final



# ---------------------------------------
# Splits: Write Splits to Excel
# ---------------------------------------
def writeExcelFileByRep(owner_value, output_folder):


    owner = str(owner_value)

    # Filter by owner
    df_abridged = df[df['Sales Representative']==owner]
    rows_target = dataframe_to_rows(df_abridged)

    # ------------
    # Write to Excel
    # ------------

    FILE_PATH = output_folder
    print(FILE_PATH)


    book = load_workbook(FILE_PATH)
    writer = ExcelWriter(FILE_PATH, engine='openpyxl')

    writer.book = book

    for sheet in book.worksheets:
        if sheet.title == 'Contacts':

            for row in sheet['A1:H4']:
               for cell in row:
                   cell.value = None

            # Replenish
            for r_idx, row in enumerate(rows_target, 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)


    constant_tries = 2000
    tries = 2000

    assert tries > 0
    error = None
    result = None

    while tries:
        try:
            writer.save()
            writer.close()
        except IOError as e:
            error = e
            tries -= 1
            print('Attempt #', (constant_tries-tries)+1)
        except ValueError as e:
            error = e
            tries -= 1
            print('Attempt #', (constant_tries-tries)+1)
        else:
            break
    if not tries:
        print('Attempt #', (constant_tries-tries)+1)
        raise error

    print('Attempt #', (constant_tries-tries)+1)
    #print(df_abridged.loc[:,'Company':'Industry'].head(5))
    print("Done writing Excel file!")




# ---------------------------------------
# Creates files per included Sales Representative
# ---------------------------------------
def loopRosterCreateFiles(reps):

    print("Creating files!")
    print(reps)

    count = 1

    for rep in reps:
        print('------')

        print(count, rep)

        count = count + 1

        output_rep = "Contacts List - " + str(rep)
        output_folder = OUTPUT_DIRECTORY+ output_rep

        # Creates Folder for each rep
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        print("Created folder for", rep)

        # Set File Names
        rep_excel_file = QUARTER+ " Contact List - "+str(rep)+ " " + VERSION +".xlsx".strip()
        rep_excel_file_no_ext = QUARTER+ " Contact List - "+str(rep)+ " " + VERSION.strip()
        rep_excel_path = output_folder+"/"+rep_excel_file

        # Copies template
        shutil.copy(FILE_TEMPLATE,rep_excel_path)

        # Writes Filtered Data to Excel
        writeExcelFileByRep(rep,rep_excel_path)

        # Uploads to Google Drive
        #loopGSpreadsheet(writeToGDrive(rep_excel_file_no_ext,rep_excel_path,getFolder(rep)))


def generateNewFolders(reps):

    # FOLDERS LOOKUP
    count = 1
    folders = {}
    managers =[]
    for rep in reps:
        foldername = "Contacts List - " + str(rep)
        if folderInGDrive(foldername) == False:
            folder_id = createGDriveFolder(foldername,PARENT_FOLDER)
            print("New Folder", foldername, folder_id)
            count += 1
        # else:
        #     print(foldername, "exists!")
        time.sleep(5)
    print("Created", count, "new folders")



def main():
    #generateNewFolders(getSalesRep())
    loopRosterCreateFiles(getSalesRep())

if __name__ == '__main__':
    main()
